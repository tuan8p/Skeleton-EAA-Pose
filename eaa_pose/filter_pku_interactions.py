"""
eaa_pose.filter_pku_interactions
=================================
Module 1 — PKU-MMD v1 daily-action filter.

Strategy
--------
Read ``Actions_v2.xlsx`` and keep only rows whose cells are highlighted
(background fill).  Those rows define the daily (non-interaction) label IDs.

For each label file in ``label_dir``:
  - Keep only lines whose ``label_id`` is in the kept set.
  - If at least one line remains, write the filtered label to
    ``out_label_dir`` and copy the matching skeleton file to
    ``out_skeleton_dir``.

Outputs
-------
1. Filtered label ``.txt`` files in ``out_label_dir``.
2. Matching skeleton ``.txt`` files in ``out_skeleton_dir``.
3. ``Actions_daily.csv`` — ``Label,Action`` for all kept label IDs.

CLI usage
---------
    python -m eaa_pose.filter_pku_interactions --config configs/pku_v1.yaml
"""

from __future__ import annotations

import argparse
import csv
import shutil
from pathlib import Path

import openpyxl

from .config import PipelineConfig


# ---------------------------------------------------------------------------
# Actions_v2.xlsx reader (colored rows)
# ---------------------------------------------------------------------------

class ColoredRowActionsReader:
    """Read kept action labels from an Actions_v2.xlsx file.

    A row is considered **kept** when at least one cell in the Label or
    Action column has a non-default background fill (the Excel highlight
    used to mark daily actions in PKU-MMD v2).
    """

    @staticmethod
    def read_kept_actions(path: Path) -> dict[int, str]:
        """Return ``{label_id: action_name}`` for highlighted rows only.

        Parameters
        ----------
        path:
            Path to ``Actions_v2.xlsx``.

        Returns
        -------
        Mapping of integer label IDs to action name strings, sorted by ID
        when written downstream.
        """
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        kept: dict[int, str] = {}

        for row_idx in range(2, ws.max_row + 1):
            label_cell = ws.cell(row_idx, 1)
            action_cell = ws.cell(row_idx, 2)

            if not (
                ColoredRowActionsReader._cell_is_highlighted(label_cell)
                or ColoredRowActionsReader._cell_is_highlighted(action_cell)
            ):
                continue

            label_id = label_cell.value
            action_name = action_cell.value
            if label_id is None or action_name is None:
                continue

            try:
                kept[int(label_id)] = str(action_name)
            except (ValueError, TypeError):
                continue

        wb.close()
        return kept

    @staticmethod
    def _cell_is_highlighted(cell) -> bool:
        """Return True if the cell has a visible background fill."""
        fill = cell.fill
        if fill is None:
            return False

        fill_type = fill.fill_type or fill.patternType
        if not fill_type or fill_type == "none":
            return False

        fg = fill.fgColor
        if fg is None:
            return False

        if fg.type == "rgb":
            rgb = (fg.rgb or "").upper()
            # Transparent / no-fill markers in Excel XML
            return rgb not in ("", "00000000", "FFFFFFFF", "00FFFFFF")

        if fg.type == "theme":
            return True

        if fg.type == "indexed":
            return fg.indexed not in (None, 0)

        return fill_type == "solid"


# ---------------------------------------------------------------------------
# Actions_daily.csv writer
# ---------------------------------------------------------------------------

class ActionsDailyCsvWriter:
    """Write the filtered action list as a CSV file."""

    @staticmethod
    def write(path: Path, actions: dict[int, str]) -> None:
        """Write ``Label,Action`` rows sorted by label ID.

        Parameters
        ----------
        path:
            Destination CSV path (e.g. ``out/Actions_daily.csv``).
        actions:
            Mapping ``{label_id: action_name}`` to write.
        """
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(["Label", "Action"])
            for label_id in sorted(actions):
                writer.writerow([label_id, actions[label_id]])


# ---------------------------------------------------------------------------
# Label file filter
# ---------------------------------------------------------------------------

def filter_label_file(label_path: Path, kept_label_ids: set[int]) -> list[str]:
    """Return label lines whose ``label_id`` is in ``kept_label_ids``.

    PKU label format per line::

        label_id,start_frame,end_frame,confidence
    """
    kept_lines: list[str] = []
    with label_path.open("r", encoding="utf-8") as fh:
        for line in fh:
            stripped = line.strip()
            if not stripped:
                continue
            parts = stripped.split(",")
            if not parts:
                continue
            try:
                label_id = int(parts[0])
            except ValueError:
                continue
            if label_id in kept_label_ids:
                kept_lines.append(stripped)
    return kept_lines


# ---------------------------------------------------------------------------
# Main filter class
# ---------------------------------------------------------------------------

class PKUDailyActionFilter:
    """Filter PKU-MMD v1 labels/skeletons by highlighted rows in Actions_v2.xlsx.

    Reads the ``filter`` section from a :class:`~eaa_pose.config.PipelineConfig`
    and performs:

    1. Parse ``Actions_v2.xlsx`` for highlighted (daily) action rows.
    2. Filter each label file, keeping only daily action instances.
    3. Copy filtered labels and matching skeleton files to ``out/``.
    4. Write ``Actions_daily.csv``.
    """

    def __init__(self, cfg: PipelineConfig) -> None:
        self._cfg = cfg

    def run(self) -> None:
        """Execute the full filter pipeline."""
        cfg = self._cfg

        skeleton_dir   = Path(cfg["filter.skeleton_dir"])
        label_dir      = Path(cfg["filter.label_dir"])
        src_actions    = Path(cfg["filter.src_actions_xlsx"])
        out_label_dir  = Path(cfg["filter.out_label_dir"])
        out_skel_dir   = Path(cfg["filter.out_skeleton_dir"])
        out_actions_csv = Path(cfg["filter.out_actions_csv"])

        self._assert_dir(skeleton_dir, "skeleton_dir")
        self._assert_dir(label_dir, "label_dir")
        self._assert_file(src_actions, "src_actions_xlsx")

        out_label_dir.mkdir(parents=True, exist_ok=True)
        out_skel_dir.mkdir(parents=True, exist_ok=True)

        kept_actions = ColoredRowActionsReader.read_kept_actions(src_actions)
        kept_label_ids = set(kept_actions)
        print(f"Kept {len(kept_actions)} daily actions from {src_actions}")
        print(f"  Label IDs: {sorted(kept_label_ids)}")

        n_label_written = 0
        n_skel_copied = 0
        n_skipped_no_kept = 0
        n_skipped_no_skel = 0

        for label_path in sorted(label_dir.glob("*.txt")):
            filtered_lines = filter_label_file(label_path, kept_label_ids)
            if not filtered_lines:
                n_skipped_no_kept += 1
                continue

            out_label_path = out_label_dir / label_path.name
            out_label_path.write_text(
                "\n".join(filtered_lines) + "\n",
                encoding="utf-8",
            )
            n_label_written += 1

            skel_path = skeleton_dir / label_path.name
            if skel_path.is_file():
                shutil.copy2(skel_path, out_skel_dir / skel_path.name)
                n_skel_copied += 1
            else:
                n_skipped_no_skel += 1

        ActionsDailyCsvWriter.write(out_actions_csv, kept_actions)

        print("\n=== Filter summary ===")
        print(f"  Label files scanned      : {len(list(label_dir.glob('*.txt')))}")
        print(f"  Label files written      : {n_label_written}")
        print(f"  Skeleton files copied    : {n_skel_copied}")
        print(f"  Skipped (no daily label) : {n_skipped_no_kept}")
        print(f"  Skipped (no skeleton)    : {n_skipped_no_skel}")
        print(f"  Filtered label dir       : {out_label_dir}")
        print(f"  Filtered skeleton dir    : {out_skel_dir}")
        print(f"  Actions_daily.csv        : {out_actions_csv}")

    @staticmethod
    def _assert_dir(path: Path, name: str) -> None:
        if not path.is_dir():
            raise FileNotFoundError(
                f"Config key '{name}' points to a missing directory: {path}"
            )

    @staticmethod
    def _assert_file(path: Path, name: str) -> None:
        if not path.is_file():
            raise FileNotFoundError(
                f"Config key '{name}' points to a missing file: {path}"
            )


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "Filter PKU-MMD v1 labels by highlighted rows in Actions_v2.xlsx."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument(
        "--config", required=True,
        help="Path to the dataset config YAML (e.g. configs/pku_v1.yaml).",
    )
    p.add_argument("--skeleton-dir",      dest="skeleton_dir",      default=None)
    p.add_argument("--label-dir",         dest="label_dir",         default=None)
    p.add_argument("--src-actions-xlsx",  dest="src_actions_xlsx",  default=None)
    p.add_argument("--out-label-dir",     dest="out_label_dir",     default=None)
    p.add_argument("--out-skeleton-dir",  dest="out_skeleton_dir",  default=None)
    p.add_argument("--out-actions-csv",   dest="out_actions_csv",   default=None)
    return p


def main() -> None:
    parser = _build_parser()
    args = parser.parse_args()
    cfg = PipelineConfig.load(args.config, cli_args=args)
    PKUDailyActionFilter(cfg).run()


if __name__ == "__main__":
    main()
